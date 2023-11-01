from openpyxl import Workbook, load_workbook

fileToReadEnteriesFrom = "book1.xlsx"
fileToWriteData = "DATA_1.xlsx"

##excel_file = "book1.xlsx"
##workbook = load_workbook(excel_file)
##worksheet = workbook.active

def readCell(fileName, cellvalue):
    excel_file = fileName
    workbook = load_workbook(excel_file)
    worksheet = workbook.active
    value = worksheet[cellvalue].value
##    print(value)
    return value


def readMaxCol(fileName):
    excel_file = fileName
    workbook = load_workbook(excel_file)
    worksheet = workbook.active
    maxNumberOfColumns = worksheet.max_column
    return maxNumberOfColumns

def readMaxRow(fileName):
    excel_file = fileName
    workbook = load_workbook(excel_file)
    worksheet = workbook.active
    maxNumberOfRow = worksheet.max_row
    return maxNumberOfRow
    
def writeToCellOfEXCEL(fileName, row, col, Value):
    excel_file = fileName
    workbook = load_workbook(excel_file)
    worksheet = workbook.active
    cell = str(col)+str(row)
    worksheet[cell] = Value
    workbook.save(excel_file)
    
    return True



### Write data to the worksheet
##for x in range(1, worksheet.max_row+1):
##    hyper = worksheet.cell(row = x, column = 4). hyperlink.display
##    cell = "F"+str(x)
##    worksheet[cell] = hyper
##
### Save the workbook to the Excel file
##workbook.save(excel_file)
